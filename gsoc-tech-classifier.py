import requests
import openpyxl
import bs4
import argparse
from fake_useragent import UserAgent

ua = UserAgent() 
header = { 
    "User-Agent": ua.random 
     } 

url = "https://summerofcode.withgoogle.com/archive/2019/organizations/"

res = requests.get(url)
res.raise_for_status()

ap = argparse.ArgumentParser()
ap.add_argument("-t","--tech",required=True,help="Technology for which Organizations to be searched")
args = vars(ap.parse_args())

language = args["lang"]
str(language)

soup = bs4.BeautifulSoup(res.text,'html.parser')

org = soup.select('h4[class="organization-card__name font-black-54"]')

orgLink = soup.find_all("a",class_="organization-card__link")
languageCheck = ['no']*len(org)
orgURL = ['none']*len(orgLink)

item = 0
wb = openpyxl.Workbook()
sheet = wb['Sheet']

sheet.cell(row=1,column=1).value="Organization"
sheet.cell(row=1,column=2).value="Does Technology Offered?"
sheet.cell(row=1,column=3).value="Link for the Organization"

for link in orgLink:
    
    presentLink = link.get('href')
    url2 = "https://summerofcode.withgoogle.com" + presentLink

    print(item)
    print(url2)

    orgURL[item] = url2
    res2 = requests.get(url2)
    res2.raise_for_status()

    soup2 = bs4.BeautifulSoup(res2.text,'html.parser')

    compTech = soup2.find_all("li",class_="organization__tag organization__tag--technology")

    for name in compTech:

        if language in name.getText():
            languageCheck[item] = 'yes'

    item = item + 1



for i in range(0,len(org)):
    sheet.cell(row = i + 2, column = 1).value= org[i].getText()
    sheet.cell(row = i + 2, column = 2).value = languageCheck[i] 
    sheet.cell(row = i + 2, column = 3).value = orgURL[i]

wb.save("gsocOrgList.xlsx")
